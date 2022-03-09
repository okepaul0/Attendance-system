from calendar import FRIDAY, MONDAY, SATURDAY, SUNDAY, THURSDAY, TUESDAY, WEDNESDAY
from openpyxl import load_workbook
book = load_workbook('attend.xlsx')
sheet = book.active

date = input('Input name in format dd/mm/yyyy:')
val = input('Choose day of the week:\n[1] SUNDAY\n[2] MONDAY\n[3] TUESDAY\n[4] WEDNESDAY\n[5] THURSDAY\n[6] FRIDAY\n[7] SATURDAY:\t')
week = ['SUNDAY','MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY','SATURDAY']

for row in  sheet['b1':'d1']:
  for index, cell in enumerate(row):
    if cell.coordinate == 'D1':
      cell.value = week[int(val)-1]
    if cell.coordinate == 'B1' :
      cell.value = date
chat = 'E'
while chat != 'E' or chat != 'e' :
  
  password = [1,2,3,4,5,6,7,8,9]
  i= 1    
  for row in  sheet['B3':'B34']:
    for index, cell in enumerate(row):
      if sheet.cell(row = int(cell.coordinate[1:]),column =3).value != '√':
        print(f"{str(int(cell.coordinate[1:])-2)} {cell.value}\n")
        i+=1
  print('\n[E] Exit')

  chat = input('Select name:\t')
  if chat == 'e' or chat =='E' :
    print('Workbook Updated')
    break
  for column in sheet['C3':'C34']:
    for index, cell in enumerate(column):
      if cell.coordinate == 'C'+str(int(chat)+2):
        cell.value = '√'
book.save('attends.xlsx')
 